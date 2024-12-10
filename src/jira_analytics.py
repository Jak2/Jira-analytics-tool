import getpass
from jira import JIRA
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def get_jira_connection():
    jira_url = input("Enter your Jira URL: ")
    username = input("Enter your Jira username: ")
    api_token = getpass.getpass("Enter your Jira API token: ")
    
    return JIRA(server=jira_url, basic_auth=(username, api_token))

def get_jql_query():
    return input("Enter your JQL query: ")

def get_fields_to_include(jira):
    all_fields = jira.fields()
    field_choices = input("Do you want to include all fields? (y/n): ").lower()
    
    if field_choices == 'y' or field_choices == 'yes':
        return [field['name'] for field in all_fields]
    else:
        selected_fields = []
        for field in all_fields:
            include = input(f"Include field '{field['name']}'? (y/n): ").lower()
            if include == 'y' or include == 'yes':
                selected_fields.append(field['name'])
        return selected_fields

def get_max_results():
    max_results = input("Enter the number of results to fetch (leave blank for all): ")
    return int(max_results) if max_results else None

def fetch_issues(jira, jql_query, fields_to_include, max_results):
    return jira.search_issues(jql_query, fields=fields_to_include, maxResults=max_results)

def save_to_excel(issues, fields_to_include, filename="jira_issues.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Jira Issues"

    # Write headers
    for col, field in enumerate(fields_to_include, start=1):
        ws.cell(row=1, column=col, value=field)

    # Write data
    for row, issue in enumerate(issues, start=2):
        for col, field in enumerate(fields_to_include, start=1):
            value = getattr(issue.fields, field, "N/A")
            ws.cell(row=row, column=col, value=str(value))

    # Adjust column widths
    for col in range(1, len(fields_to_include) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(filename)
    print(f"Results saved to {filename}")

def main():
    jira = get_jira_connection()
    jql_query = get_jql_query()
    fields_to_include = get_fields_to_include(jira)
    max_results = get_max_results()
    
    issues = fetch_issues(jira, jql_query, fields_to_include, max_results)
    save_to_excel(issues, fields_to_include)

if __name__ == "__main__":
    main()